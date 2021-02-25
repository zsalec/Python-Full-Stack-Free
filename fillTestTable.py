"""
# 将测评试题转换的导入系统的 xls 格式
# @ version: 0.0.1
# @ Author: Alec
# @ Create Time: 2021-02-24 15:32:53
# @ Modified by: Alec
# @ Modified time: 2021-02-24 15:33:15
# @ Description:
# @ 将测评试题转换的导入系统的 xls 格式 (从文本文件中读取数据)
"""

from typing import List
from openpyxl import Workbook, load_workbook
from datetime import date
import os

def parseFile(filenameEvaluation: str, filenameDimension: str, minScore: int):
    '''
    解析文件
    '''
    with open(filenameEvaluation, 'r', encoding='utf-8') as f:
        data = f.readlines()
        dimensions = loadDimension(filenameDimension)
        questions = parseData(data, dimensions)
        fileDim = '维度_' + date.today().strftime('%Y%m%d') + '.xlsx'
        saveDimension(dimensions, fileDim)
        fileQue = '试题_' + date.today().strftime('%Y%m%d') + '.xlsx'
        saveQuestions(questions, fileQue)


def parseData(lines: List, dimensions: dict):
    no = 1
    # dimensions = dict()
    questions = []
    rowNo = 0
    while rowNo < len(lines):
        line = lines[rowNo]
        rowNo += 1
        errCode, questionInfo = parseQuestion(line, no, dimensions)
        if errCode == 0:
            no += 1
            questions.append(questionInfo)
            rowTemp = rowNo
            while rowTemp < len(lines):
                line = lines[rowTemp]
                rowTemp += 1
                errCode, answers = parseAnswer(line, questionInfo['orientation'], minScore)
                if errCode == 0:
                    questionInfo['answers'] = answers
                    break
            rowNo = rowTemp
    print(questions)
    return questions


def parseQuestion(line: str, no: int, dimension: dict):
    '''
    分析题目
    '''
    errCode = 1
    questionInfo = {}
    line = line.strip()
    if line[0] == '-':
        orientation = False
        line = line[1:]
    else:
        orientation = True
    pos = line.find('、')
    # items = line.split('、')
    # if len(items) == 2:
    if pos > 0:
        if line[:pos].strip() == f'{no}':
            # obj = re.match('\((.*)\)', items[1])
            # if obj == None:
            #     obj = re.match('（(.*)）', items[1])
            # print('match object', obj)
            errCode = 0
            title, dims = parseDimension(line[pos+1:])
            print(title, dims)
            dimsNo = []
            for dim in dims:
                if dim not in dimension.keys():
                    dimension[dim] = 101 + len(dimension)
                dimsNo.append({'no': dimension[dim], 'name': dim})
            questionInfo = {'no': no, 'title': title, 'orientation': orientation, 'dimensions': dimsNo}
    return errCode, questionInfo


def parseDimension(line: str):
    '''
    分析维度和题干
    '''
    title = line.strip()
    dimension = []
    while True:
        if title[-1] == '）' or title[-1] == ')':
            pos = title.rfind('(')
            if pos < 0:
                pos = title.rfind('（')
            if pos > 0:
                dimension.append(title[pos+1:-1])
                title = title[:pos]
                continue
        break
    return title, dimension


def clearAnswerData(answer: str):
    '''
    清洗答案数据
    '''
    makes = ['A ', 'a ', 'B ', 'b ', 'C ', 'c ', 'D ', 'd ', 'E ', 'e ']
    s = answer.strip()
    for make in makes:
        while s.find(make) >= 0:
            s = s.replace(make, make.strip())
    items = s.split(' ')
    return items


def parseAnswer(line: str, orientation: bool, minScore: int):
    '''
    分析答案
    '''
    answerInfo = []
    errCode = 1
    #line = clearAnswerData(line.strip())
    #items = line.split(' ')
    items = clearAnswerData(line)
    if len(items) > 0:
        errCode = 0
        print(items)
        for item in items:
            if len(item) > 0 and item[0] == chr(65 + len(answerInfo)):
                answerInfo.append({'mark': chr(65 + len(answerInfo)), 'title': item[1:], 'score': minScore + len(answerInfo)})
        if orientation:
            for aItem in answerInfo:
                aItem['score'] = len(answerInfo) + minScore - aItem['score']
    return errCode, answerInfo

def saveDimension(dimensions: List, filename: str):
    '''
    保存维度到 Excel 文件中
    '''
    wb = Workbook()
    ws = wb.active
    ws.title = 'Dimensions'
    # 填写标题
    titles = ['维度ID', '名称']
    # for i in range(len(titles)):
    #     ws[f'{chr(65+i)}1'] = titles[i]
    fillTitle(ws, 1, titles)
    # 填写维度信息
    row = 2
    for key in dimensions:
        # ws[f'A{row}'] = dimensions[key]
        # ws[f'B{row}'] = key
        data = {
            '维度ID': dimensions[key],
            '名称': key
        }
        fillRowData(ws, row, data, titles)
        row += 1
    wb.save(filename)
    wb.close()

def saveQuestions(questions: List, filename: str):
    '''
    保存试题到 Excel 文件中
    '''
    # 编号	题目类型	问题描述	A选项	A分值	B选项	B分值	C选项	C分值	D选项	D分值	E选项	E分值	试题非试题	维度	备注
    # 填写标题
    wb = Workbook()
    ws = wb.active
    ws.title = 'Questions'
    titles = ['编号', '题目类型', '问题描述', 'A选项', 'A分值', 'B选项', 'B分值', 'C选项', 'C分值', 'D选项', 'D分值', 'E选项', 'E分值', '试题非试题', '维度', '备注']
    # for i in range(len(titles)):
    #     ws[f'{chr(65+i)}1'] = titles[i]
    fillTitle(ws, 1, titles)
    row = 2
    for i in range(len(questions)):
        data = {
            '编号': i + 1,
            '题目类型': '选择题',
            '问题描述': questions[i]['title'],
            '试题非试题': '是',
            '备注': i + 1
        }
        # 填写得分
        for j in range(5):
            ans = ''
            score = ''
            if j < len(questions[i]['answers']):
                ans = questions[i]['answers'][j]['title']
                score = questions[i]['answers'][j]['score']
            data[f'{chr(65+j)}选项'] = ans
            data[f'{chr(65+j)}分值'] = score
        # 填写维度
        dim = ''
        for j in questions[i]['dimensions']:
            dim += f"{j['no']},"
        data['维度'] = dim.strip(',')
        fillRowData(ws, row, data, titles)
        row += 1
    wb.save(filename)
    wb.close()

def fillTitle(ws, row: int, titles: List):
    '''
    填写一行数据
    '''
    for i in range(len(titles)):
        ws[f'{chr(65+i)}{row}'] = titles[i]

def fillRowData(ws, row: int, data: dict, titles: List):
    '''
    填写一行数据
    '''
    for i in range(len(titles)):
        ws[f'{chr(65+i)}{row}'] = data[titles[i]]


def loadDimension(filename: str):
    '''
    加载以往维度信息
    '''
    dimensions = {}
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            no = ws.cell(row, 1).value
            if isinstance(no, int):
                name = ws.cell(row, 2).value
                dimensions[name] = no
    return dimensions


if __name__ == '__main__':
    # str = '1、我认为自己是个坚强的人。（能力）（效能）'
    # dimension = {}
    # no = 1
    # errCode, questionInfo = parseQuestion(str, no, dimension)
    # print('test quesntion info', questionInfo, dimension)

    # answer = 'A完全符合   B大部分符合  C不太符合  D完全不符合'
    # errCode, answerInfo = parseAnswer(answer, True, 1)
    # print('test answer info', answerInfo)
    # filename = 'E:\\MZWB\\中学版\\EvaluationInfo.txt'
    filename = 'E:\\MZWB\\学习压力测评中学版（单页版）\\Evaluation.txt'
    filenameDim = '维度_20210225.xlsx'
    minScore = 1
    parseFile(filenameEvaluation=filename, minScore=minScore, filenameDimension=filenameDim)
    print('Done!!!')
